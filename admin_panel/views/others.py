import json
from copy import copy

from django.core import serializers
from django.core.exceptions import ObjectDoesNotExist
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.urls import reverse_lazy, reverse
from django.views.generic import *
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import range_boundaries

from admin_panel.forms import *
from admin_panel.models import *


def statistic(request):
    return render(request, 'admin_panel/statistic.html')


class PayboxList(ListView):
    template_name = 'admin_panel/paybox.html'
    context_object_name = 'paybox'
    queryset = Paybox.objects.all()

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        total_plus = sum(Paybox.objects.filter(debit_credit='plus', is_complete=True).values_list('total', flat=True))
        total_minus = sum(Paybox.objects.filter(debit_credit='minus', is_complete=True).values_list('total', flat=True))

        context['total_plus'] = total_plus
        context['total_minus'] = total_minus
        return context


class CreatePaybox(FormView):
    def get(self, request, income, *args, **kwargs):
        form = PayboxForm()
        form.fields['date_published'].initial = timezone.now().date()

        if income == 'plus':
            form.fields['article'].queryset = Article.objects.filter(debit_credit="plus")
        elif income == 'minus':
            form.fields['article'].queryset = Article.objects.filter(debit_credit="minus")
        data = {
            'income': income,
            'form': form,
        }
        return render(request, 'admin_panel/get_paybox_form.html', context=data)

    def post(self, request, income, *args, **kwargs):
        form = PayboxForm(request.POST)

        if form.is_valid():
            instance = form.save()
            if income == 'plus':
                if instance.personal_account is not None and instance.is_complete is True:
                    personal_account = PersonalAccount.objects.get(pk=instance.personal_account_id)
                    personal_account.balance = personal_account.balance + instance.total
                    personal_account.save()
                instance.debit_credit = 'plus'
            elif income == 'minus':
                instance.debit_credit = 'minus'
            instance.save()
            return redirect('paybox')
        else:

            data = {
                'income': income,
                'form': form,
            }
            return render(request, 'admin_panel/get_paybox_form.html', context=data)


class CopyPaybox(CreatePaybox):
    def get(self, request, pk, *args, **kwargs):
        copy = Paybox.objects.get(pk=pk)
        form = PayboxForm(instance=copy)

        if copy.debit_credit == 'plus':
            form.fields['article'].queryset = Article.objects.filter(debit_credit="plus")
        elif copy.debit_credit == 'minus':
            form.fields['article'].queryset = Article.objects.filter(debit_credit="minus")
        income = copy.debit_credit
        data = {
            'income': income,
            'form': form,
        }
        return render(request, 'admin_panel/get_paybox_form.html', context=data)


from urllib.parse import urlencode, quote


class PayboxDetail(DetailView):
    model = Paybox
    template_name = 'admin_panel/read_paybox.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        paybox = Paybox.objects.get(pk=self.kwargs['pk'])
        context['paybox'] = paybox

        wb = Workbook()
        ws = wb.active  # это лист в excel
        ws.append(['Платёж', f'#{paybox.number}'])
        ws.append(['Дата', f'{paybox.date_published}'])
        if paybox.flat_owner:
            ws.append(['Владелец квартиры',
                       f'{paybox.flat_owner.user.last_name} {paybox.flat_owner.user.first_name} {paybox.flat_owner.patronymic}'])
        else:
            ws.append(['Владелец квартиры', f'не указан'])

        if paybox.personal_account:
            ws.append(['Лицевой счёт', paybox.personal_account.number])
        else:
            ws.append(['Лицевой счёт', f'не указан'])

        ws.append(['Приход/Расход', paybox.get_debit_credit_display()])
        ws.append(['Проведён', 'Проведён' if paybox.is_complete else 'Не проведён'])
        if paybox.article:
            ws.append(['Статья', paybox.article.title])
        else:
            ws.append(['Статья', f'не указана'])

        ws.append(['Квитанция', ''])
        ws.append(['Услуга', ''])
        ws.append(['Сумма', f'{paybox.total}'])
        ws.append(['Валюта', 'UAH'])
        ws.append(['Комментарий', paybox.comment])
        ws.append(['Приход/Расход', paybox.get_debit_credit_display()])
        ws.append(['Проведён', 'Проведён' if paybox.is_complete else 'Не проведён'])
        if paybox.user:
            ws.append(['Менеджер', f'{paybox.user.user.last_name} {paybox.user.user.first_name}'])
        else:
            ws.append(['Менеджер', f'не указан'])
        ws.title = "Выписка"  # это название листа в excel
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40

        wb.save('media/paybox/info.xlsx')

        # Get full path to workbook
        wb.close()

        return context


class UpdatePaybox(FormView):
    def get(self, request, pk, *args, **kwargs):
        paybox = Paybox.objects.get(pk=pk)
        form = PayboxForm(instance=paybox)
        if paybox.debit_credit == 'plus':
            form.fields['article'].queryset = Article.objects.filter(debit_credit="plus")
        elif paybox.debit_credit == 'minus':
            form.fields['article'].queryset = Article.objects.filter(debit_credit="minus")
        data = {
            'form': form,
        }
        return render(request, 'admin_panel/update_paybox.html', context=data)

    def post(self, request, pk, *args, **kwargs):
        paybox = Paybox.objects.get(pk=pk)
        form = PayboxForm(request.POST, instance=paybox)
        if form.is_valid():
            instance = form.save()
            if instance.debit_credit == 'plus':
                if instance.personal_account is not None:
                    personal_account = PersonalAccount.objects.get(pk=instance.personal_account_id)
                    total = sum(
                        Paybox.objects.filter(personal_account=personal_account,
                                              is_complete=True).values_list('total', flat=True))
                    personal_account.balance = total
                    personal_account.save()
            return redirect('paybox')
        else:
            data = {
                'form': form,
            }
            return render(request, 'admin_panel/update_paybox.html', context=data)


class DeletePaybox(FormView):
    def post(self, request, pk, *args, **kwargs):
        paybox = Paybox.objects.get(pk=pk)
        if paybox.debit_credit == 'plus':
            if paybox.personal_account is not None and paybox.is_complete is True:
                personal_account = PersonalAccount.objects.get(pk=paybox.personal_account_id)
                paybox.delete()
                total = sum(
                    Paybox.objects.filter(personal_account=personal_account,
                                          is_complete=True).values_list('total', flat=True))
                personal_account.balance = total
                personal_account.save()
            else:
                paybox.delete()

        else:
            paybox.delete()
        return redirect('paybox')


class ReceiptList(ListView):
    template_name = 'admin_panel/receipts.html'
    context_object_name = 'receipts'
    queryset = Receipt.objects.all()


class CreateReceipt(CreateView):
    model = Receipt
    template_name = 'admin_panel/get_receipt_form.html'
    form_class = ReceiptForm
    success_url = reverse_lazy('receipts')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['indications'] = Indication.objects.order_by('date_published').all()
        context['service_formset'] = ReceiptServiceFormset(queryset=ReceiptService.objects.none(), prefix='service')
        context['measures'] = Measure.objects.all()
        context['services'] = Service.objects.all()
        return context

    def post(self, request, *args, **kwargs):
        service_formset = ReceiptServiceFormset(request.POST, prefix='service')
        receipt_form = ReceiptForm(request.POST)
        if receipt_form.is_valid() and service_formset.is_valid():
            obj = receipt_form.save()
            instances = service_formset.save(commit=False)
            for instance in instances:
                instance.receipt_id = obj.id
                instance.save()
            return redirect('receipts')
        else:
            data = {
                "indications": Indication.objects.order_by('date_published').all(),
                "service_formset": service_formset,
                "form": receipt_form,
                'measures': Measure.objects.all(),
                'services': Service.objects.all(),
            }
            return render(request, 'admin_panel/get_receipt_form.html', context=data)


class GetIndicationsSortedList(View):
    def get(self, request, flat_id):
        indications = Indication.objects.order_by('date_published').filter(flat_id=flat_id)
        data = {
            "indications": indications,
        }
        return render(request, 'admin_panel/ajax_indication_table.html', context=data)


class UpdateReceipt(UpdateView):
    model = Receipt
    template_name = 'admin_panel/update_receipt.html'
    form_class = ReceiptForm
    success_url = reverse_lazy('receipts')

    def get(self, request, *args, **kwargs):
        receipt = Receipt.objects.get(id=self.kwargs['pk'])
        data = {
            'indications': Indication.objects.order_by('date_published').filter(flat_id=receipt.flat.id),
            'form': ReceiptForm(instance=receipt,
                                initial={'house': receipt.flat.house, 'section': receipt.flat.section}),
            'service_formset': ReceiptServiceFormset(prefix='service', queryset=ReceiptService.objects.filter(
                receipt_id=self.kwargs['pk'])),
            'services': Service.objects.all(),
            'measures': Measure.objects.all(),
        }
        return render(request, 'admin_panel/update_receipt.html', context=data)

    def post(self, request, pk, *args, **kwargs):
        service_formset = ReceiptServiceFormset(request.POST, prefix='service')
        receipt_form = ReceiptForm(request.POST, instance=Receipt.objects.get(id=pk))

        if receipt_form.is_valid() and service_formset.is_valid():
            obj = receipt_form.save()
            instances = service_formset.save()
            print(service_formset.errors)
            print(instances)
            for instance in instances:
                print('hello')
                instance.receipt_id = obj.id
                instance.save()
            return redirect('receipts')
        else:
            data = {
                "indications": Indication.objects.order_by('date_published').all(),
                "service_formset": service_formset,
                "form": receipt_form,
                'measures': Measure.objects.all(),
                'services': Service.objects.all(),
            }
            return render(request, 'admin_panel/update_receipt.html', context=data)


class ReceiptPrint(View):
    def get(self, request, pk, *args, **kwargs):
        receipt = Receipt.objects.get(pk=pk)
        rows = ReceiptExcelDoc.objects.all().order_by('id')

        data = {
            'rows': rows,
            'receipt': receipt
        }
        return render(request, 'admin_panel/receipt_printing.html', context=data)


class ReceiptDownloadExcel(View):
    def post(self, request, excel_id, receipt_id, *args, **kwargs):
        receipt = Receipt.objects.get(pk=receipt_id)
        services = ReceiptService.objects.filter(receipt_id=receipt_id)

        doc = ReceiptExcelDoc.objects.get(pk=excel_id)
        wb = load_workbook(doc.file)
        ws = wb.active  # это лист в excel
        for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=25):
            for cell in row:
                match cell.value:
                    case 'personal_accountNumber':
                        cell.value = receipt.flat.personal_account.number
                        # cell.alignment = Alignment(horizontal='center', vertical='center') # выровнять по центру
                    case 'personalManager':
                        cell.value = str(receipt.flat.flat_owner)
                        cell.alignment = Alignment(horizontal='center', vertical='center')  # выровнять по центру
                    case 'receiptNumber':
                        cell.value = receipt.number
                    case 'receiptStartDate':
                        cell.value = str(receipt.start_date)
                    case 'totalPrice':
                        cell.value = receipt.total_price
                    case 'flatOwner':
                        cell.value = str(receipt.flat.flat_owner)
                    case 'personalAccountBalance':
                        cell.value = receipt.flat.personal_account.balance
                    case 'receiptDatePublished':
                        cell.value = receipt.date_published
                        cell.alignment = Alignment(horizontal='center', vertical='center')  # выровнять по центр
                    case 'receiptMonthPublished':
                        month_dict = {
                            'January': 'Январь',
                            'February': 'Февраль',
                            'March': 'Март',
                            'April': 'Апрель',
                            'May': 'Май',
                            'June': 'Июнь',
                            'July': 'Июль',
                            'August': 'Август',
                            'September': 'Сентябрь',
                            'October': 'Октябрь',
                            'November': 'Ноябрь',
                            'December': 'Декабрь'
                        }
                        cell.value = month_dict[str(receipt.date_published.strftime('%B'))]
                        cell.alignment = Alignment(horizontal='center', vertical='center')  # выровнять по центр
                    case 'total':
                        cell.value = receipt.total_price
                    case 'service':
                        pass


                    #     for i, obj in enumerate(services):
                    #         ws.insert_rows(cell.row + i)
                    #         ws.cell(row=cell.row + i, column=cell.column, value=str(obj.service.title))
                    # case 'tariff':
                    #     for i, obj in enumerate(services):
                    #         ws.cell(row=cell.row + i, column=cell.column, value=str(obj.receipt.tariff.title))
                    # case 'measure':
                    #     for i, obj in enumerate(services):
                    #         ws.cell(row=cell.row + i, column=cell.column, value=str(obj.measure.title))
                    # case 'totalServicePrice':
                    #     for i, obj in enumerate(services):
                    #         ws.cell(row=cell.row + i, column=cell.column, value=str(float(obj.consumption) * float(obj.unit_price)))
        ws.title = "Квитанция"  # это название листа в excel
        wb.save('media/receipt/result/report.xlsx')

        data = {
        }
        return HttpResponse(json.dumps(data), content_type='application/json')


class ReceiptPrintingSettings(View):
    def get(self, request, *args, **kwargs):
        rows = ReceiptExcelDoc.objects.all().order_by('id')
        form = ReceiptExcelDocForm()
        data = {
            'rows': rows,
            'form': form,
        }
        return render(request, 'admin_panel/receipt_printing_settings.html', context=data)

    def post(self, request, *args, **kwargs):
        form = ReceiptExcelDocForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
        return redirect('receipt_print_settings')


class ReceiptPrintingSettingsDefault(View):
    def get(self, request, pk, *args, **kwargs):
        for row in ReceiptExcelDoc.objects.all():
            row.by_default = False
            row.save()
        row = ReceiptExcelDoc.objects.get(pk=pk)
        row.by_default = True
        row.save()
        return redirect('receipt_print_settings')


class ReceiptPrintingSettingsDelete(DeleteView):
    success_url = reverse_lazy('receipt_print_settings')
    queryset = ReceiptExcelDoc.objects.all()


class CopyReceipt(FormView):
    def get(self, request, pk, *args, **kwargs):
        copy = Receipt.objects.get(pk=pk)
        receipt_form = ReceiptForm(instance=copy, initial={'house': copy.flat.house, 'section': copy.flat.section})

        service_formset = ReceiptServiceFormset(queryset=ReceiptService.objects.filter(receipt_id=pk),
                                                prefix='service')
        data = {
            'form': receipt_form,
            'service_formset': service_formset,
            "indications": Indication.objects.order_by('date_published').filter(flat_id=copy.flat.id),
            'measures': Measure.objects.all(),
            'services': Service.objects.all(),
        }
        return render(request, 'admin_panel/get_receipt_form.html', context=data)

    def post(self, request, *args, **kwargs):
        service_formset = ReceiptServiceFormset(request.POST, prefix='service')
        receipt_form = ReceiptForm(request.POST)
        if receipt_form.is_valid() and service_formset.is_valid():
            obj = receipt_form.save()
            instances = service_formset.save(commit=False)
            for instance in instances:
                instance.receipt_id = obj.id
                instance.save()
            return redirect('receipts')
        else:
            data = {
                "indications": Indication.objects.order_by('date_published').all(),
                "service_formset": service_formset,
                "form": receipt_form,
                'measures': Measure.objects.all(),
                'services': Service.objects.all(),
            }
            return render(request, 'admin_panel/get_receipt_form.html', context=data)


class ReceiptDetail(DetailView):
    model = Receipt
    template_name = 'admin_panel/read_receipt.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        receipt = Receipt.objects.get(pk=self.kwargs['pk'])
        receipt_services = ReceiptService.objects.filter(receipt_id=self.kwargs['pk'])
        context['receipt'] = receipt
        context['receipt_services'] = receipt_services
        return context


class DeleteReceipt(DeleteView):
    success_url = reverse_lazy('receipts')
    queryset = Receipt.objects.all()


class FlatListView(ListView):
    template_name = 'admin_panel/flats.html'
    context_object_name = 'flats'
    queryset = Flat.objects.all()


class CreateFlatView(CreateView):
    model = Flat
    template_name = 'admin_panel/get_flat_form.html'
    form_class = FlatForm
    success_url = reverse_lazy('flats')


class FlatDetail(DetailView):
    model = Flat
    template_name = 'admin_panel/read_flat.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        flat = Flat.objects.get(pk=self.kwargs['pk'])
        context['flat'] = flat
        return context


class UpdateFlatView(UpdateView):
    model = Flat
    template_name = 'admin_panel/update_flat.html'
    form_class = FlatForm
    success_url = reverse_lazy('flats')


class DeleteFlatView(DeleteView):
    success_url = reverse_lazy('flats')
    queryset = Flat.objects.all()


class FlatAcceptPayment(CreatePaybox):
    def get(self, request, pk, *args, **kwargs):
        initial = {}
        flat = Flat.objects.get(pk=pk)
        if flat.personal_account is not None and flat.flat_owner is not None:
            initial['flat_owner'] = flat.flat_owner
            initial['personal_account'] = flat.personal_account

        form = PayboxForm(initial=initial)
        form.fields['article'].queryset = Article.objects.filter(debit_credit="plus")
        form.fields['date_published'].initial = timezone.now().date()

        data = {
            'income': 'plus',
            'form': form,
        }
        return render(request, 'admin_panel/get_paybox_form.html', context=data)


class FlatAcceptReceipt(FormView):
    def get(self, request, pk, *args, **kwargs):
        flat = Flat.objects.get(pk=pk)

        receipt_form = ReceiptForm(
            initial={
                'house': flat.house,
                'section': flat.section,
                'flat': flat,
                'tariff': flat.tariff,
                'personal_account': flat.personal_account,
            }
        )

        service_formset = ReceiptServiceFormset(queryset=ReceiptService.objects.none(),
                                                prefix='service')
        data = {
            'form': receipt_form,
            'service_formset': service_formset,
            "indications": Indication.objects.order_by('date_published').filter(flat_id=flat.id),
            'measures': Measure.objects.all(),
            'services': Service.objects.all(),
        }
        return render(request, 'admin_panel/get_receipt_form.html', context=data)

    def post(self, request, *args, **kwargs):
        service_formset = ReceiptServiceFormset(request.POST, prefix='service')
        receipt_form = ReceiptForm(request.POST)
        if receipt_form.is_valid() and service_formset.is_valid():
            obj = receipt_form.save()
            instances = service_formset.save(commit=False)
            for instance in instances:
                instance.receipt_id = obj.id
                instance.save()
            return redirect('receipts')
        else:
            data = {
                "indications": Indication.objects.order_by('date_published').all(),
                "service_formset": service_formset,
                "form": receipt_form,
                'measures': Measure.objects.all(),
                'services': Service.objects.all(),
            }
            return render(request, 'admin_panel/get_receipt_form.html', context=data)


class FlatReceiptList(ListView):
    template_name = 'admin_panel/receipts.html'
    context_object_name = 'receipts'

    def get_queryset(self):
        flat = Flat.objects.get(pk=self.kwargs['pk'])
        queryset = Receipt.objects.filter(flat_id=flat.id)
        return queryset


class FlatPayboxList(ListView):
    template_name = 'admin_panel/paybox.html'
    context_object_name = 'paybox'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        flat = Flat.objects.get(pk=self.kwargs['pk'])
        if flat.personal_account is not None:
            total_plus = sum(Paybox.objects.filter(debit_credit='plus',
                                                   personal_account=flat.personal_account,
                                                   is_complete=True).values_list('total', flat=True))
        else:
            total_plus = 0
        context['total_plus'] = total_plus
        return context

    def get_queryset(self):
        flat = Flat.objects.get(pk=self.kwargs['pk'])
        if flat.personal_account is not None:
            queryset = Paybox.objects.filter(personal_account=flat.personal_account)
        else:
            queryset = Paybox.objects.none()

        return queryset


class PersonalAccountListView(ListView):
    template_name = 'admin_panel/personal_accounts.html'
    context_object_name = 'personal_accounts'
    queryset = PersonalAccount.objects.all()


class PersonalAccountDetail(DetailView):
    model = PersonalAccount
    template_name = 'admin_panel/read_personal_account.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        personal_account = PersonalAccount.objects.get(pk=self.kwargs['pk'])
        context['personal_account'] = personal_account
        return context


class CreatePersonalAccount(CreateView):
    model = PersonalAccount
    template_name = 'admin_panel/get_personal_account_form.html'
    form_class = PersonalAccountForm
    success_url = reverse_lazy('personal_accounts')


class UpdatePersonalAccount(UpdateView):
    model = PersonalAccount
    template_name = 'admin_panel/update_personal_account.html'
    form_class = PersonalAccountForm
    success_url = reverse_lazy('personal_accounts')


class DeletePersonalAccount(DeleteView):
    success_url = reverse_lazy('personal_accounts')
    queryset = PersonalAccount.objects.all()


class PersonalAccountAcceptPayment(CreatePaybox):
    def get(self, request, pk, *args, **kwargs):
        initial = {}
        personal_account = PersonalAccount.objects.get(pk=pk)
        initial['personal_account'] = personal_account
        if personal_account.flat is not None and personal_account.flat.flat_owner is not None:
            initial['flat_owner'] = personal_account.flat.flat_owner
        form = PayboxForm(initial=initial)
        form.fields['article'].queryset = Article.objects.filter(debit_credit="plus")
        form.fields['date_published'].initial = timezone.now().date()

        data = {
            'income': 'plus',
            'form': form,
        }
        return render(request, 'admin_panel/get_paybox_form.html', context=data)


class PersonalAccountAcceptReceipt(FormView):
    def get(self, request, pk, *args, **kwargs):
        personal_account = PersonalAccount.objects.get(pk=pk)

        receipt_form = ReceiptForm(
            initial={
                'house': personal_account.flat.house,
                'section': personal_account.flat.section,
                'flat': personal_account.flat,
                'tariff': personal_account.flat.tariff,
                'personal_account': personal_account,
            }
        )

        service_formset = ReceiptServiceFormset(queryset=ReceiptService.objects.none(),
                                                prefix='service')
        data = {
            'form': receipt_form,
            'service_formset': service_formset,
            "indications": Indication.objects.order_by('date_published').filter(flat_id=personal_account.flat.id),
            'measures': Measure.objects.all(),
            'services': Service.objects.all(),
        }
        return render(request, 'admin_panel/get_receipt_form.html', context=data)

    def post(self, request, *args, **kwargs):
        service_formset = ReceiptServiceFormset(request.POST, prefix='service')
        receipt_form = ReceiptForm(request.POST)
        if receipt_form.is_valid() and service_formset.is_valid():
            obj = receipt_form.save()
            instances = service_formset.save(commit=False)
            for instance in instances:
                instance.receipt_id = obj.id
                instance.save()
            return redirect('receipts')
        else:
            data = {
                "indications": Indication.objects.order_by('date_published').all(),
                "service_formset": service_formset,
                "form": receipt_form,
                'measures': Measure.objects.all(),
                'services': Service.objects.all(),
            }
            return render(request, 'admin_panel/get_receipt_form.html', context=data)


class PersonalAccountReceiptList(ListView):
    template_name = 'admin_panel/receipts.html'
    context_object_name = 'receipts'

    def get_queryset(self):
        personal_account = PersonalAccount.objects.get(pk=self.kwargs['pk'])
        if personal_account.flat is not None:
            queryset = Receipt.objects.filter(flat_id=personal_account.flat.id)
        else:
            queryset = Receipt.objects.none()
        return queryset


class PersonalAccountPayboxList(ListView):
    template_name = 'admin_panel/paybox.html'
    context_object_name = 'paybox'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        total_plus = sum(Paybox.objects.filter(debit_credit='plus',
                                               personal_account_id=self.kwargs['pk'],
                                               is_complete=True).values_list('total', flat=True))

        context['total_plus'] = total_plus
        return context

    def get_queryset(self):
        personal_account = PersonalAccount.objects.get(pk=self.kwargs['pk'])
        queryset = Paybox.objects.filter(personal_account=personal_account)
        return queryset


class ClientListView(ListView):
    template_name = 'admin_panel/clients.html'
    context_object_name = 'clients'
    queryset = FlatOwner.objects.all()


class ClientDetail(DetailView):
    model = FlatOwner
    template_name = 'admin_panel/read_client.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        client = FlatOwner.objects.prefetch_related("flat_set").get(pk=self.kwargs['pk'])

        context['client'] = client
        context['flats'] = client.flat_set.all()

        return context


class ClientSignUpView(CreateView):
    model = CustomUser
    form_class = ClientSignUpForm
    template_name = 'admin_panel/get_client_form.html'

    def get_context_data(self, **kwargs):
        return super().get_context_data(**kwargs)

    def form_valid(self, form):
        form.save()
        return redirect('clients')


class UpdateClientView(UpdateView):
    model = CustomUser
    form_class = ClientUpdateForm
    template_name = 'admin_panel/update_client.html'
    success_url = reverse_lazy('clients')
    queryset = CustomUser.objects.all()

    def get(self, request, pk, *args, **kwargs):
        user = CustomUser.objects.get(pk=pk)
        owner = FlatOwner.objects.get(user=user)
        form = ClientUpdateForm(instance=user,
                                initial={'birthday': owner.birthday, 'viber': owner.viber, 'telegram': owner.telegram,
                                         'patronymic': owner.patronymic, 'ID': owner.ID,
                                         'bio': owner.bio, })
        data = {
            'form': form,
        }
        return render(request, 'admin_panel/update_client.html', context=data)


class DeleteClientView(DeleteView):
    success_url = reverse_lazy('clients')
    queryset = CustomUser.objects.all()


class HouseListView(ListView):
    template_name = 'admin_panel/houses.html'
    context_object_name = 'houses'
    queryset = House.objects.all()


class HouseDetail(DetailView):
    model = House
    template_name = 'admin_panel/read_house.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        house = House.objects.prefetch_related('gallery__photo_set', 'houseuser_set').get(pk=self.kwargs['pk'])
        photos = house.gallery.photo_set.all()
        users = house.houseuser_set.all()
        context['house'] = house
        context['photos'] = photos
        context['users'] = users

        return context


class CreateHouseView(FormView):
    def get(self, request, *args, **kwargs):
        house_form = HouseForm()
        personals = Personal.objects.all()
        photo_formset = HousePhotoFormset(queryset=Photo.objects.none(), prefix='gallery')
        section_formset = SectionFormset(queryset=Section.objects.none(), prefix='section')
        floor_formset = FloorFormset(queryset=Floor.objects.none(), prefix='floor')
        house_user_formset = HouseUserFormset(queryset=HouseUser.objects.none(), prefix='personal')
        data = {
            'house_form': house_form,
            'personals': personals,
            'photo_formset': photo_formset,
            'section_formset': section_formset,
            'floor_formset': floor_formset,
            'house_user_formset': house_user_formset,
        }
        return render(request, 'admin_panel/get_house_form.html', context=data)

    def post(self, request, *args, **kwargs):

        house_form = HouseForm(request.POST, request.FILES)
        section_formset = SectionFormset(request.POST, request.FILES, prefix='section')
        floor_formset = FloorFormset(request.POST, request.FILES, prefix='floor')
        photo_formset = HousePhotoFormset(request.POST, request.FILES, prefix='gallery')
        house_user_formset = HouseUserFormset(request.POST, request.FILES, prefix='personal')

        if section_formset.is_valid() and floor_formset.is_valid() and house_form.is_valid() and photo_formset.is_valid() and house_user_formset.is_valid():
            house_obj = house_form.save()
            house_user_instances = house_user_formset.save(commit=False)
            for instance in house_user_instances:
                instance.house_id = house_obj.id
                instance.save()
            gallery = Gallery.objects.create()
            house_obj.gallery_id = gallery.id
            for photo_form in photo_formset:
                instance = photo_form.save(commit=False)
                instance.gallery_id = gallery.id
                instance.save()
            section_instances = section_formset.save(commit=False)
            floor_instances = floor_formset.save(commit=False)
            for instance in section_instances:
                instance.house_id = house_obj.id

                instance.save()
            for instance in floor_instances:
                instance.house_id = house_obj.id
                instance.save()
            house_obj.save()

        else:
            data = {
                'house_form': house_form,
                'section_formset': section_formset,
                'floor_formset': floor_formset,
                'photo_formset': photo_formset,
                'house_user_formset': house_user_formset,
            }
            return render(request, 'admin_panel/get_house_form.html', context=data)
        return redirect('houses')


class UpdateHouseView(FormView):
    def get(self, request, pk, *args, **kwargs):
        personals = Personal.objects.all()
        house = House.objects.get(pk=pk)
        house_form = HouseForm(instance=house)
        photo_formset = PhotoFormset(queryset=Photo.objects.filter(gallery_id=house.gallery_id), prefix='gallery')
        section_formset = SectionFormset(queryset=Section.objects.filter(house=house), prefix='section')
        floor_formset = FloorFormset(queryset=Floor.objects.filter(house=house), prefix='floor')
        house_user_formset = HouseUserFormset(queryset=HouseUser.objects.filter(house_id=house.id), prefix='personal')
        data = {
            'house_form': house_form,
            'personals': personals,
            'house': house,
            'photo_formset': photo_formset,
            'section_formset': section_formset,
            'floor_formset': floor_formset,
            'house_user_formset': house_user_formset,
        }
        return render(request, 'admin_panel/update_house.html', context=data)

    def post(self, request, pk, *args, **kwargs):
        house = House.objects.get(pk=pk)
        house_form = HouseForm(request.POST, request.FILES, instance=house)
        section_formset = SectionFormset(request.POST, request.FILES, prefix='section')
        floor_formset = FloorFormset(request.POST, request.FILES, prefix='floor')
        photo_formset = PhotoFormset(request.POST, request.FILES, prefix='gallery')
        house_user_formset = HouseUserFormset(request.POST, request.FILES, prefix='personal')

        if section_formset.is_valid() and floor_formset.is_valid() and house_form.is_valid() and photo_formset.is_valid() and house_user_formset.is_valid():
            house_obj = house_form.save()
            house_user_instances = house_user_formset.save(commit=False)
            for instance in house_user_instances:
                instance.house_id = house_obj.id
                instance.save()
            photo_instances = photo_formset.save(commit=False)
            gallery = Gallery.objects.get(house=house)
            for photo_instance in photo_instances:
                photo_instance.gallery_id = gallery.id
                photo_instance.save()

            section_instances = section_formset.save(commit=False)
            floor_instances = floor_formset.save(commit=False)
            for instance in section_instances:
                instance.house_id = house_obj.id
                instance.save()
            for instance in floor_instances:
                instance.house_id = house_obj.id
                instance.save()
            house_obj.save()

        else:
            data = {
                'house_form': house_form,
                'section_formset': section_formset,
                'floor_formset': floor_formset,
                'photo_formset': photo_formset,
                'house_user_formset': house_user_formset,
            }
            return render(request, 'admin_panel/get_house_form.html', context=data)
        return redirect('houses')


class GetRoleView(View):
    def get(self, request, pk):
        personal = Personal.objects.get(pk=pk)
        data = {
            'role': personal.get_role_display(),
        }

        return HttpResponse(json.dumps(data), content_type='application/json')


class GetSectionInfoView(View):
    def get(self, request, pk):
        section = Section.objects.prefetch_related('flat_set').get(pk=pk)
        flats = serializers.serialize('json', section.flat_set.all())
        data = {
            "flats": flats,
        }
        return JsonResponse(data, safe=False)


class GetFlatInfoView(View):
    def get(self, request, pk):
        data = {}
        flat = Flat.objects.get(pk=pk)
        try:
            flat_owner_obj = FlatOwner.objects.get(flat=flat)
            flat_owner = serializers.serialize('json', [flat_owner_obj])
            data['flat_owner'] = flat_owner
            user = CustomUser.objects.get(flatowner=flat_owner_obj)
            user = serializers.serialize('json', [user])
            data['user'] = user
        except ObjectDoesNotExist:
            pass
        try:
            if flat.personal_account is not None:
                personal_account = flat.personal_account
                personal_account = serializers.serialize('json', [personal_account])
                data['personal_account'] = personal_account
        except ObjectDoesNotExist:
            pass
        try:
            if flat.tariff is not None:
                tariff = TariffSystem.objects.get(pk=flat.tariff.pk)
                tariff = serializers.serialize('json', [tariff])
                data['tariff'] = tariff
        except ObjectDoesNotExist:
            pass
        return JsonResponse(data, safe=False)


class GetHouseInfoView(View):
    def get(self, request, pk):
        house = House.objects.prefetch_related('section_set', 'floor_set', 'flat_set').get(pk=pk)
        sections = serializers.serialize('json', house.section_set.all())
        floors = serializers.serialize('json', house.floor_set.all())
        flats = serializers.serialize('json', house.flat_set.all())
        house = serializers.serialize('json', [house])
        data = {
            "house": house,
            "sections": sections,
            "floors": floors,
            "flats": flats,
        }
        return JsonResponse(data, safe=False)


class GetTariffInfoView(View):
    def get(self, request, pk):
        tariff = TariffSystem.objects.prefetch_related('tariffservice_set').get(pk=pk)
        tariff_services = serializers.serialize('json', tariff.tariffservice_set.all())
        data = {
            "tariff_services": tariff_services,
        }
        return JsonResponse(data, safe=False)


class GetServiceInfoView(View):
    def get(self, request, pk):
        service = Service.objects.get(pk=pk)
        service = serializers.serialize('json', [service])
        data = {
            "service": service,
        }
        return JsonResponse(data, safe=False)


class GetIndicationInfoView(View):
    def get(self, request, flat_id, service_id):
        indication = Indication.objects.filter(flat_id=flat_id, service_id=service_id)
        if indication.count() != 0:
            indication = serializers.serialize('json', indication)
            data = {
                "indication": indication,
            }
        else:
            return JsonResponse({}, safe=False)

        return JsonResponse(data, safe=False)


class GetFlatOwnerInfo(View):
    def get(self, request, pk):
        flat_owner = FlatOwner.objects.prefetch_related('flat_set').get(pk=pk)
        flats = flat_owner.flat_set.select_related('house').all()
        personal_accounts = PersonalAccount.objects.filter(flat__in=flats.values_list("id", flat=True))
        flats = serializers.serialize('json', flats)
        personal_accounts = serializers.serialize('json', personal_accounts)
        data = {
            "flats": flats,
            "personal_accounts": personal_accounts,
        }
        return JsonResponse(data, safe=False)


class GetAllFlats(View):
    def get(self, request):
        flats = Flat.objects.all()
        flats = serializers.serialize('json', flats)
        all_personal_accounts = PersonalAccount.objects.all()
        all_personal_accounts = serializers.serialize('json', all_personal_accounts)

        data = {
            "flats": flats,
            "all_personal_accounts": all_personal_accounts,
        }
        return JsonResponse(data, safe=False)


class DeleteHouseView(DeleteView):
    def post(self, request, pk, *args, **kwargs):
        house = House.objects.get(pk=pk)
        gallery = Gallery.objects.get(pk=house.gallery_id)
        house.delete()
        gallery.delete()
        return redirect('houses')


class MailboxList(ListView):
    template_name = 'admin_panel/mailbox.html'
    context_object_name = 'mailboxes'
    queryset = MailBox.objects.all()


class CreateMailbox(CreateView):
    model = MailBox
    template_name = 'admin_panel/get_mailbox_form.html'
    form_class = MailBoxForm
    success_url = reverse_lazy('mailboxes')


class MailboxDetail(DetailView):
    model = MailBox
    template_name = 'admin_panel/read_mailbox.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['mailbox'] = MailBox.objects.get(pk=self.kwargs['pk'])
        return context


class DeleteMailbox(DeleteView):
    success_url = reverse_lazy('mailboxes')
    queryset = MailBox.objects.all()


class ApplicationList(ListView):
    template_name = 'admin_panel/applications.html'
    context_object_name = 'applications'
    queryset = Application.objects.all()


class ApplicationDetail(DetailView):
    model = Application
    template_name = 'admin_panel/read_application.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        application = Application.objects.get(pk=self.kwargs['pk'])
        context['application'] = application
        return context


class CreateApplication(CreateView):
    model = Application
    template_name = 'admin_panel/get_application_form.html'
    form_class = CreateApplicationForm
    success_url = reverse_lazy('applications')


class UpdateApplication(UpdateView):
    model = Application
    template_name = 'admin_panel/update_application.html'
    form_class = ApplicationForm
    success_url = reverse_lazy('applications')


class DeleteApplication(DeleteView):
    success_url = reverse_lazy('applications')
    queryset = Application.objects.all()


class CounterList(ListView):
    template_name = 'admin_panel/counters.html'
    context_object_name = 'indications'
    queryset = Indication.objects.order_by('flat', 'service', '-date_published').distinct('flat', 'service')


class CounterIndicationsList(ListView):
    template_name = 'admin_panel/counter_indications_list.html'
    context_object_name = 'indications'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['flat'] = Flat.objects.get(pk=self.kwargs['flat'])
        context['service'] = Service.objects.get(pk=self.kwargs['service'])
        return context

    def get_queryset(self):
        queryset = Indication.objects.filter(flat_id=self.kwargs['flat'], service_id=self.kwargs['service'])
        return queryset


class FlatIndicationsList(ListView):
    template_name = 'admin_panel/flat_indications_list.html'
    context_object_name = 'indications'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['flat'] = Flat.objects.get(pk=self.kwargs['flat'])
        return context

    def get_queryset(self):
        queryset = Indication.objects.filter(flat_id=self.kwargs['flat'])
        return queryset


class IndicationDetail(DetailView):
    model = Indication
    template_name = 'admin_panel/read_indication.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        indication = Indication.objects.get(pk=self.kwargs['pk'])
        context['indication'] = indication
        return context


class CreateIndication(CreateView):
    model = Indication
    template_name = 'admin_panel/get_indication_form.html'
    form_class = IndicationForm
    success_url = reverse_lazy('counters')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['create_new_indication'] = False
        return context

    def post(self, request, *args, **kwargs):
        indication_form = IndicationForm(request.POST)
        if indication_form.is_valid():
            obj = indication_form.save()
            if request.POST['action'] == 'save_and_new':
                indication_form = IndicationForm(initial={'indication_val': obj.indication_val, 'service': obj.service})
                data = {
                    'form': indication_form
                }
                return render(request, 'admin_panel/get_indication_form.html', context=data)
            else:
                return redirect('counters')
        else:
            data = {
                'form': indication_form
            }
            return render(request, 'admin_panel/get_indication_form.html', context=data)


class CreateNewIndication(CreateView):
    model = Indication
    template_name = 'admin_panel/get_indication_form.html'
    form_class = IndicationForm

    def get_success_url(self):
        category = self.kwargs['category']  # Retrieve the category value from the URL
        return '/category/{}/'.format(category)  # Build the URL using the retrieved value

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['create_new_indication'] = True
        return context

    def get_initial(self):
        initial = super().get_initial()
        flat = Flat.objects.get(pk=self.kwargs['flat'])
        service = Service.objects.get(pk=self.kwargs['service'])
        initial['house'] = flat.house
        initial['section'] = flat.section
        initial['flat'] = flat
        initial['service'] = service
        return initial

    def post(self, request, *args, **kwargs):
        form = IndicationForm(request.POST)
        if form.is_valid():
            obj = form.save()
            if request.POST['action'] == 'save_and_new':
                indication_form = IndicationForm(initial={'indication_val': obj.indication_val, 'service': obj.service})
                data = {
                    'form': indication_form
                }
                return render(request, 'admin_panel/get_indication_form.html', context=data)
            else:
                url = reverse('counter_indications', args=[obj.flat.id, obj.service.id])
                return redirect(url)
        else:
            data = {
                'form': form,
                'create_new_indication': True,
            }
            return render(request, 'admin_panel/get_indication_form.html', context=data)


class CreateIndicationForFlat(CreateView):
    model = Indication
    template_name = 'admin_panel/get_indication_form.html'
    form_class = IndicationForm

    def get_success_url(self):
        category = self.kwargs['category']  # Retrieve the category value from the URL
        return '/category/{}/'.format(category)  # Build the URL using the retrieved value

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['create_new_indication'] = True
        return context

    def get_initial(self):
        initial = super().get_initial()
        flat = Flat.objects.get(pk=self.kwargs['flat'])
        initial['house'] = flat.house
        initial['section'] = flat.section
        initial['flat'] = flat
        return initial

    def post(self, request, *args, **kwargs):
        form = IndicationForm(request.POST)
        if form.is_valid():
            obj = form.save()
            if request.POST['action'] == 'save_and_new':
                indication_form = IndicationForm(initial={'indication_val': obj.indication_val, 'service': obj.service})
                data = {
                    'form': indication_form
                }
                return render(request, 'admin_panel/get_indication_form.html', context=data)
            else:
                url = reverse('flat_indications', args=[obj.flat.id])
                return redirect(url)
        else:
            data = {
                'form': form,
                'create_new_indication': True,
            }
            return render(request, 'admin_panel/get_indication_form.html', context=data)


class UpdateIndication(UpdateView):
    model = Indication
    template_name = 'admin_panel/update_indication.html'
    form_class = IndicationForm
    success_url = reverse_lazy('counters')

    def get(self, request, pk, *args, **kwargs):
        obj = Indication.objects.get(pk=pk)
        form = IndicationForm(instance=obj, initial={'house': obj.flat.house, 'section': obj.flat.section, })
        data = {
            'form': form,
        }
        return render(request, 'admin_panel/update_indication.html', context=data)


class DeleteIndication(DeleteView):
    success_url = reverse_lazy('counters')
    queryset = Indication.objects.all()


from openpyxl import Workbook

# class DownloadExcel(View):
#     def get(self, request, pk, *args, **kwargs):
#         paybox = Paybox.objects.get(pk=pk)
#         wb = Workbook()
#         ws = wb.active  # это лист в excel
#         ws.title = "Выписка1"# это название листа в excel
#         wb.save('balances.xlsx')
#         data = {
#         }
#         return redirect(request.path)
